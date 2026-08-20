import datetime
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.download_hire_report import (
    IMPORT_COLUMNS,
    IMPORT_TABLE_NAME,
    IMPORT_WORKBOOK_NAME,
    IMPORT_WORKSHEET_NAME,
    calculate_nco_window,
    create_excel_import,
    transform_symplr_data,
)


class SymplrV2ImportTests(unittest.TestCase):
    def test_nco_window_examples(self):
        cases = [
            (
                datetime.date(2026, 8, 19),
                {
                    "previous_nco": datetime.date(2026, 8, 17),
                    "nco_1": datetime.date(2026, 8, 31),
                    "nco_2": datetime.date(2026, 9, 14),
                    "nco_3": datetime.date(2026, 9, 28),
                    "extraction_start": datetime.date(2026, 8, 17),
                    "extraction_end": datetime.date(2026, 9, 28),
                },
            ),
            (
                datetime.date(2026, 9, 2),
                {
                    "previous_nco": datetime.date(2026, 8, 31),
                    "nco_1": datetime.date(2026, 9, 14),
                    "nco_2": datetime.date(2026, 9, 28),
                    "nco_3": datetime.date(2026, 10, 12),
                    "extraction_start": datetime.date(2026, 8, 31),
                    "extraction_end": datetime.date(2026, 10, 12),
                },
            ),
        ]

        for run_date, expected in cases:
            with self.subTest(run_date=run_date):
                window = calculate_nco_window(today=run_date)
                for key, value in expected.items():
                    self.assertEqual(window[key], value)

    def test_transform_outputs_only_symplr_owned_sharepoint_columns(self):
        raw = pd.DataFrame(
            [
                {
                    "Applicant Name": "Jane Example",
                    "Email": "  JANE.EXAMPLE@EMAIL.COM ",
                    "Job Title": "RN",
                    "Job Code": "RN001",
                    "Hired Date": "8/1/2026 10:30:00 AM",
                    "Start Date": "08/31/2026",
                    "Phone": "312-555-1000",
                    "Recruiter": "Recruiter A",
                    "Hiring Manager": "Manager A",
                    "Account": "Sinai",
                    "Facility": "Hospital",
                    "Facility Code": "FAC1",
                    "Department": "Dept A",
                    "Department Code": "2370",
                    "Orientation 1 Date": "08/31/2026",
                },
                {
                    "Applicant Name": "Blank Email",
                    "Email": " ",
                    "Job Title": "RN",
                    "Job Code": "RN001",
                    "Hired Date": "8/1/2026",
                    "Start Date": "08/31/2026",
                    "Phone": "",
                    "Recruiter": "",
                    "Hiring Manager": "",
                    "Account": "",
                    "Facility": "",
                    "Facility Code": "",
                    "Department": "",
                    "Department Code": "",
                    "Orientation 1 Date": "08/31/2026",
                },
            ]
        )

        output, stats = transform_symplr_data(raw)

        self.assertEqual(list(output.columns), IMPORT_COLUMNS)
        self.assertEqual(len(output), 1)
        self.assertEqual(output.loc[0, "Email"], "jane.example@email.com")
        self.assertEqual(output.loc[0, "Start date"], "08/31/2026")
        self.assertEqual(output.loc[0, "NCO date"], "08/31/2026")
        self.assertNotIn("Hiring Manager Email", output.columns)
        self.assertEqual(stats["blank_email_rows_rejected"], 1)

    def test_duplicate_email_keeps_latest_current_record(self):
        raw = pd.DataFrame(
            [
                {
                    "Applicant Name": "Person Example",
                    "Email": "person@example.com",
                    "Job Title": "Old Title",
                    "Job Code": "OLD",
                    "Hired Date": "08/01/2026",
                    "Start Date": "08/31/2026",
                    "Phone": "111",
                    "Recruiter": "Recruiter A",
                    "Hiring Manager": "Manager A",
                    "Account": "Sinai",
                    "Facility": "Facility A",
                    "Facility Code": "FAC-A",
                    "Department": "Dept A",
                    "Department Code": "2370",
                    "Orientation 1 Date": "08/31/2026",
                },
                {
                    "Applicant Name": "Person Example",
                    "Email": " PERSON@EXAMPLE.COM ",
                    "Job Title": "New Title",
                    "Job Code": "NEW",
                    "Hired Date": "08/02/2026",
                    "Start Date": "09/14/2026",
                    "Phone": "222",
                    "Recruiter": "Recruiter B",
                    "Hiring Manager": "Manager B",
                    "Account": "Sinai",
                    "Facility": "Facility B",
                    "Facility Code": "FAC-B",
                    "Department": "Dept B",
                    "Department Code": "2654",
                    "Orientation 1 Date": "09/14/2026",
                },
            ]
        )

        output, stats = transform_symplr_data(raw)

        self.assertEqual(len(output), 1)
        self.assertEqual(stats["duplicate_emails_detected"], 1)
        self.assertEqual(output.loc[0, "Email"], "person@example.com")
        self.assertEqual(output.loc[0, "Job Title"], "New Title")
        self.assertEqual(output.loc[0, "Start date"], "09/14/2026")
        self.assertEqual(output.loc[0, "NCO date"], "09/14/2026")
        self.assertEqual(output.loc[0, "Hiring Manager"], "Manager B")
        self.assertEqual(output.loc[0, "Department Code"], "2654")

    def test_excel_output_contains_real_table(self):
        import_df = pd.DataFrame(
            [
                {
                    column: ""
                    for column in IMPORT_COLUMNS
                }
            ]
        )
        import_df.loc[0, "Applicant Name"] = "Jane Example"
        import_df.loc[0, "Email"] = "jane.example@email.com"
        import_df.loc[0, "Start date"] = "08/31/2026"
        import_df.loc[0, "NCO date"] = "08/31/2026"

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = create_excel_import(import_df, Path(temp_dir))

            self.assertEqual(output_path.name, IMPORT_WORKBOOK_NAME)
            workbook = load_workbook(output_path)
            self.assertIn(IMPORT_WORKSHEET_NAME, workbook.sheetnames)
            worksheet = workbook[IMPORT_WORKSHEET_NAME]
            self.assertIn(IMPORT_TABLE_NAME, worksheet.tables)
            self.assertEqual(worksheet.tables[IMPORT_TABLE_NAME].ref, "A1:O2")


if __name__ == "__main__":
    unittest.main()
